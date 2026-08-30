# Copyright (C) 2026 willywu <pop2585158@gmail.com>
# SPDX-License-Identifier: GPL-3.0-only
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""FastAPI 應用：隧道、群組視窗、錨點（含 WebSocket 廣播）、照片串流。"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from fastapi import FastAPI, HTTPException, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from PIL import Image, ImageOps
from pydantic import BaseModel, Field

from . import thumbs
from .db import Workspace
from .fsutil import platform_roots
from .interp import AnchorOrderError, AnchorRangeError
from .importer import CameraInput, ImportRequest, TunnelImporter, _ScannedPhoto
from .service import MergeConflict, TunnelService

# 輕量任務暫存（輪詢式進度）；狀態同步落地 index.db（R9），
# server 重啟後 done job 仍可由 fingerprint 復用掃描結果、running 標記 interrupted
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor

_jobs: dict[str, dict] = {}
_job_threads: dict[str, threading.Thread] = {}
_JOBS_TTL = 24 * 3600
_JOBS_MAX = 32


def _scan_to_json(scanned: list[_ScannedPhoto]) -> list[dict]:
    return [
        {
            "c": p.camera_seq,
            "p": str(p.path),
            "t": p.t.isoformat(timespec="seconds"),
            "s": p.time_source,
            "f": p.flagged,
            "w": p.width,
            "h": p.height,
            "o": p.orientation,
        }
        for p in scanned
    ]


def _scan_from_json(items: list[dict]) -> list[_ScannedPhoto]:
    from datetime import datetime as _dt

    return [
        _ScannedPhoto(
            int(i["c"]),
            Path(i["p"]),
            _dt.fromisoformat(i["t"]),
            i["s"],
            bool(i["f"]),
            i.get("w"),
            i.get("h"),
            i.get("o", 1),
        )
        for i in items
    ]


def _scan_fingerprint(body: ImportBody) -> str:
    """掃描結果只取決於機位資料夾清單（順序敏感）；commit 時比對以免誤用他人掃描。"""
    return json.dumps([c.folder for c in body.cameras], ensure_ascii=False)


class CameraBody(BaseModel):
    name: str
    folder: str
    rotation: int = 0
    grid_pos: int = -1


class ImportBody(BaseModel):
    name: str = Field(min_length=1)
    start_m: int
    end_m: int
    tolerance_seconds: float = Field(gt=0)
    layout_cols: str | int = "auto"
    cameras: list[CameraBody] = Field(min_length=1)
    project_id: int | None = None


class LayoutBody(BaseModel):
    cols: str | int


class FsRecentBody(BaseModel):
    path: str


class AnchorBody(BaseModel):
    mileage_m: int


class RealignBody(BaseModel):
    tolerance_seconds: float = Field(gt=0)


class MergeBody(BaseModel):
    direction: str
    keep: str | None = None


class CameraUpdateBody(BaseModel):
    name: str | None = None
    rotation: int | None = None
    grid_pos: int | None = None


class PhotoRotationBody(BaseModel):
    angle: int


class AnomalyItemBody(BaseModel):
    id: int | None = None
    type_id: int
    note: str | None = None


class AnnotationBody(BaseModel):
    note: str | None = None
    items: list[AnomalyItemBody] = []

class DefectTypeBody(BaseModel):
    name: str = Field(min_length=1)


class ProjectBody(BaseModel):
    name: str = Field(min_length=1)


class MoveBody(BaseModel):
    project_id: int | None = None


class GroupVisibilityBody(BaseModel):
    hidden: bool


def _read_orientation_tag(path: Path) -> int:
    """讀 EXIF orientation(274)；讀取失敗視為 1（不轉正）。"""
    try:
        with Image.open(path) as im:
            return int((im.getexif() or {}).get(274, 1) or 1)
    except Exception:
        return 1


def _request_etags(request: Request) -> set[str]:
    """解析 If-None-Match（可為逗號清單或 *；* 以空集合回表示全部符合由呼叫端判斷）。"""
    raw = request.headers.get("if-none-match")
    if not raw:
        return set()
    raw = raw.strip()
    if raw == "*":
        return {"*"}
    return {t.strip().removeprefix("W/") for t in raw.split(",") if t.strip()}


def _delete_thumbs_for(root: Path, tid: int, photo_ids: list[int]) -> None:
    """刪除指定照片的所有縮圖快取（任何寬度/版本）。"""
    if not photo_ids:
        return
    cache_dir = root / ".thumb_cache"
    if not cache_dir.is_dir():
        return
    prefixes = tuple(f"{tid}_{pid}_" for pid in photo_ids)
    for f in cache_dir.iterdir():
        if f.is_file() and f.name.startswith(prefixes):
            f.unlink(missing_ok=True)


def _bump_pixels(service: TunnelService, workspace: Workspace, tid: int, photo_ids: list[int]) -> None:
    """像素版本遞增＋舊縮圖清除——僅真正改變像素的操作（R9 精準失效）。"""
    if not photo_ids:
        return
    service.bump_pixel_versions(tid, photo_ids)
    _delete_thumbs_for(workspace.root, tid, photo_ids)


def _pregen_thumbs(workspace: Workspace, tid: int) -> None:
    """commit 後背景預生成 w=1600 縮圖（R9）：把冷啟風暴變暖快取。

    與請求內生成共用 thumbs.single_flight；失敗靜默（留待請求時重試）。
    """
    try:
        conn = workspace.open_tunnel(tid)
        rows = conn.execute(
            "SELECT p.id AS pid, c.root_path AS root, p.rel_path AS rel, "
            "COALESCE(p.rotation_override, -1) AS rov, COALESCE(c.rotation, 0) AS crot, "
            "COALESCE(p.pixel_version, 0) AS pv, p.orientation AS orient "
            "FROM photos p JOIN cameras c ON c.id = p.camera_id "
            # orientation NULL（舊隧道未 backfill）先跳過——避免以預設值生成
            # 錯誤方向的縮圖後被同名快取命中；留待請求時 backfill 再生成
            "WHERE p.orientation IS NOT NULL ORDER BY p.id"
        ).fetchall()
        conn.close()
        if not rows:
            return
        try:
            workers = max(1, int(os.environ.get("TUNNELVIEW_THUMB_WORKERS", "") or 4))
        except ValueError:
            workers = 4
        cache_dir = Path(workspace.root) / ".thumb_cache"

        def gen(r):
            src = Path(r["root"]) / r["rel"]
            if not src.is_file():
                return
            # 防競爭：旋轉/轉正可能已遞增版本——寫出前重查，避免復活舊版縮圖
            try:
                chk = workspace.open_tunnel(tid)
                cur = chk.execute(
                    "SELECT pixel_version FROM photos WHERE id = ?", (r["pid"],)
                ).fetchone()
                chk.close()
                if cur is None or int(cur["pixel_version"]) != int(r["pv"]):
                    return
            except Exception:
                pass
            extra = (r["rov"] if r["rov"] >= 0 else r["crot"]) % 360
            needs_t = int(r["orient"] or 1) not in (0, 1)
            variants = [(1600, extra)] + ([(1600, 0)] if extra else [])
            for w_, ex_ in variants:
                cf = cache_dir / f"{tid}_{r['pid']}_{w_}_{ex_}_v{r['pv']}.jpg"
                try:
                    thumbs.get_or_make(
                        cf,
                        lambda s=src, w_=w_, ex_=ex_, nt=needs_t: thumbs.make_thumbnail(
                            s, w_, needs_transpose=nt, extra_rotation=ex_
                        ),
                    )
                except Exception:
                    pass

        with ThreadPoolExecutor(max_workers=min(workers, len(rows))) as ex:
            list(ex.map(gen, rows))
    except Exception:
        pass


def _needs_exif_transpose(path: Path) -> bool:
    try:
        with Image.open(path) as im:
            return int((im.getexif() or {}).get(274, 1)) not in (0, 1)
    except Exception:
        return False


class _Hub:
    """每條隧道一個房間；寫入事件廣播給所有連線中的檢視端。"""

    def __init__(self):
        self._rooms: dict[int, list[WebSocket]] = defaultdict(list)
        self._loops: dict[int, asyncio.AbstractEventLoop] = {}

    async def join(self, tunnel_id: int, ws: WebSocket) -> None:
        self._loops[tunnel_id] = asyncio.get_running_loop()
        await ws.accept()
        self._rooms[tunnel_id].append(ws)

    def leave(self, tunnel_id: int, ws: WebSocket) -> None:
        try:
            self._rooms[tunnel_id].remove(ws)
        except ValueError:
            pass

    def broadcast(self, tunnel_id: int, payload: dict) -> None:
        conns = [w for w in self._rooms.get(tunnel_id, []) if w.client_state.name == "CONNECTED"]
        loop = self._loops.get(tunnel_id)
        if not conns or loop is None:
            return

        async def _send():
            for w in conns:
                try:
                    await w.send_json(payload)
                except Exception:
                    self.leave(tunnel_id, w)

        try:
            asyncio.run_coroutine_threadsafe(_send(), loop)
        except RuntimeError:
            pass


def create_app(workspace: Workspace) -> FastAPI:
    app = FastAPI(title="Tunnel View", docs_url=None, redoc_url=None)
    service = TunnelService(workspace)
    importer = TunnelImporter(workspace)
    hub = _Hub()

    # R9：啟動還原——上次執行中的 job 已無對應執行緒，誠實標記為 interrupted
    try:
        workspace.job_interrupt_running()
    except Exception:
        pass

    def _prune_jobs():
        workspace.job_prune(_JOBS_TTL, _JOBS_MAX)
        now = time.time()
        expired = [k for k, v in _jobs.items() if now - v.get("created_at", now) > _JOBS_TTL]
        for k in expired:
            _jobs.pop(k, None)

    @app.get("/api/tunnels")
    def list_tunnels():
        return workspace.list_tunnels_full()

    def _to_req(body: ImportBody) -> ImportRequest:
        return ImportRequest(
            name=body.name,
            start_m=body.start_m,
            end_m=body.end_m,
            tolerance_seconds=body.tolerance_seconds,
            layout_cols=body.layout_cols,
            cameras=[
                CameraInput(name=c.name, folder=c.folder, rotation=c.rotation, grid_pos=c.grid_pos)
                for c in body.cameras
            ],
        )

    @app.delete("/api/tunnels/{tid}")
    def remove_tunnel(tid: int):
        try:
            workspace.delete_tunnel(tid)
        except KeyError:
            raise HTTPException(404, "隧道不存在")
        return {"ok": True}

    @app.post("/api/tunnels/preview")
    def preview_import(body: ImportBody):
        # 同步 fallback（小檔量），保留舊契約
        return importer.preview(_to_req(body))

    class _JobImportBody(ImportBody):
        pass

    @app.post("/api/import/jobs/preview")
    def create_import_job(body: ImportBody):
        _prune_jobs()
        job_id = uuid.uuid4().hex[:12]
        req = _to_req(body)
        # 快速列舉不開檔即得總量
        try:
            enum_info = importer.enumerate(req)
            total = enum_info["total_valid"]
        except Exception as e:
            raise HTTPException(400, str(e))
        # 背景執行緒做 scan+align：POST 立即回 running，前端輪詢 GET 取進度；
        # 掃描結果暫存於 runtime 並落地 index.db（scan_json），commit 帶同一
        # job_id 且指紋一致時免二次 EXIF 掃描——跨 server 重啟亦然（R9）
        fp = _scan_fingerprint(body)
        _jobs[job_id] = {
            "job_id": job_id,
            "status": "running",
            "stage": "scan",
            "total": total,
            "done": 0,
            "req": req,
            "scan_fp": fp,
            "created_at": time.time(),
            "enum_info": enum_info,
        }
        try:
            workspace.job_save(job_id, status="running", stage="scan", total=total, fingerprint=fp)
        except Exception:
            pass

        def _progress(n: int):
            job = _jobs.get(job_id)
            if job is not None:
                job["done"] = min(int(job.get("done", 0)) + n, total)

        def _run():
            from dataclasses import asdict

            job = _jobs.get(job_id)
            if job is None:
                return
            try:
                scanned_photos = importer.scan(req, progress=_progress)
                preview = importer.preview(req, scanned=scanned_photos)
                if job_id not in _jobs:
                    return  # job 已被刪除：不得復活 runtime 或持久化列
                _jobs[job_id] = {
                    **job,
                    "status": "done",
                    "stage": "done",
                    "done": total,
                    "preview": asdict(preview),
                    "scan": scanned_photos,
                }
                try:
                    workspace.job_save(
                        job_id,
                        status="done",
                        stage="done",
                        done=total,
                        total=total,
                        preview_json=asdict(preview),
                        scan_json=_scan_to_json(scanned_photos),
                        fingerprint=job.get("scan_fp"),
                    )
                except Exception:
                    pass
            except Exception as e:
                cur = _jobs.get(job_id)
                if cur is not None:
                    _jobs[job_id] = {
                        "job_id": job_id,
                        "status": "failed",
                        "error": str(e),
                        "created_at": cur.get("created_at", time.time()),
                    }
                try:
                    workspace.job_save(job_id, status="failed", error=str(e))
                except Exception:
                    pass

        th = threading.Thread(target=_run, name=f"import-job-{job_id}", daemon=True)
        _job_threads[job_id] = th
        th.start()
        return {"job_id": job_id, "status": "running", "total": total, "done": 0}

    @app.get("/api/import/jobs/{job_id}")
    def get_import_job(job_id: str):
        _prune_jobs()
        job = _jobs.get(job_id)
        if job is not None:
            payload = {
                "job_id": job_id,
                "status": job.get("status"),
                "stage": job.get("stage"),
                "done": job.get("done", 0),
                "total": job.get("total", 0),
                "preview": job.get("preview"),
            }
            if job.get("error"):
                payload["error"] = job["error"]
            return payload
        row = workspace.job_get(job_id)
        if not row:
            raise HTTPException(404, "任務不存在或已逾期")
        return {
            "job_id": job_id,
            "status": row["status"],
            "stage": row.get("stage"),
            "done": row.get("done", 0),
            "total": row.get("total", 0),
            "preview": row.get("preview"),
            **({"error": row["error"]} if row.get("error") else {}),
        }

    @app.delete("/api/import/jobs/{job_id}")
    def delete_import_job(job_id: str):
        _jobs.pop(job_id, None)
        _job_threads.pop(job_id, None)
        try:
            workspace.job_delete(job_id)
        except Exception:
            pass
        return {"ok": True}

    @app.post("/api/tunnels")
    def create_tunnel(body: ImportBody, job_id: str | None = None):
        # project 歸屬先驗證（隧道建立成功後歸檔）
        if body.project_id is not None:
            names = {p["id"] for p in workspace.list_projects()}
            if body.project_id not in names:
                raise HTTPException(404, "專案不存在")
        # job_id 快路徑：掃描結果仍在暫存（或已落地 DB）且機位資料夾指紋一致時直接復用，
        # 免對（可能是慢速網路碟的）原始資料夾做第二次 EXIF 掃描
        scanned = None
        if job_id:
            th = _job_threads.get(job_id)
            if th is not None and th.is_alive():
                th.join(timeout=600)  # 防禦：前端未等 done 就送 commit 時，等背景掃描完成
            fp = _scan_fingerprint(body)
            job = _jobs.get(job_id)
            if (
                job
                and job.get("status") == "done"
                and job.get("scan") is not None
                and job.get("scan_fp") == fp
            ):
                scanned = job["scan"]
            else:
                # 重啟後 runtime 已失：從持久化的 scan_json 復用（R9）
                row = workspace.job_get(job_id)
                if (
                    row
                    and row.get("status") == "done"
                    and row.get("scan")
                    and row.get("fingerprint") == fp
                ):
                    try:
                        scanned = _scan_from_json(row["scan"])
                    except Exception:
                        scanned = None
        info = importer.commit(_to_req(body), scanned=scanned)
        for cam in body.cameras:
            workspace.add_recent_path(cam.folder)
        if body.project_id is not None:
            try:
                workspace.move_tunnel(info.tunnel_id, body.project_id)
            except Exception:
                pass  # 歸檔失敗不阻斷建立；可事後移動
        # R9：背景預生成 1600px 縮圖（可用 TUNNELVIEW_THUMB_PREGEN=0 停用）
        if os.environ.get("TUNNELVIEW_THUMB_PREGEN", "") != "0":
            threading.Thread(
                target=_pregen_thumbs,
                args=(workspace, info.tunnel_id),
                name=f"pregen-{info.tunnel_id}",
                daemon=True,
            ).start()
        # 清理對應 job
        if job_id:
            _jobs.pop(job_id, None)
            _job_threads.pop(job_id, None)
            try:
                workspace.job_delete(job_id)
            except Exception:
                pass
        return {"tunnel_id": info.tunnel_id}

    @app.get("/api/cache/scan")
    def get_scan_cache_info():
        # 簡易資訊，前端清除按鈕用
        return {"ok": True}

    @app.delete("/api/cache/scan")
    def clear_scan_cache(folder: str | None = None):
        # R9：掃描快取落地 index.db.scan_cache；此端點全清
        try:
            workspace.scan_cache_clear()
        except Exception:
            pass
        return {"ok": True}

    @app.get("/api/tunnels/{tid}/meta")
    def tunnel_meta(tid: int):
        return service.meta(tid)

    @app.put("/api/tunnels/{tid}")
    def rename_tunnel(tid: int, body: ProjectBody):
        try:
            workspace.rename_tunnel(tid, body.name)
        except KeyError:
            raise HTTPException(404, "隧道不存在")
        except ValueError as e:
            raise HTTPException(400, str(e))
        hub.broadcast(tid, {"type": "tunnel_renamed", "name": body.name.strip()})
        return {"ok": True}

    @app.get("/api/tunnels/{tid}/overview")
    def tunnel_overview(tid: int):
        return service.overview(tid)

    @app.get("/api/tunnels/{tid}/groups")
    def groups(tid: int, around: int = 0, before: int = 25, after: int = 50):
        return service.get_window(tid, around=max(around, 0), before=max(before, 0), after=max(after, 0))

    @app.get("/api/tunnels/{tid}/groups/by_mileage")
    def group_by_mileage(tid: int, m: int):
        hit = service.nearest_by_mileage(tid, m)
        if hit is None:
            raise HTTPException(404, "此隧道沒有任何群組")
        return hit

    @app.put("/api/tunnels/{tid}/groups/{seq}/visibility")
    def set_group_visibility(tid: int, seq: int, body: GroupVisibilityBody):
        try:
            service.set_group_hidden(tid, seq, body.hidden)
        except KeyError:
            raise HTTPException(404, "群組不存在")
        hub.broadcast(tid, {"type": "group_visibility", "group_seq": seq, "hidden": body.hidden})
        return {"ok": True}

    @app.get("/api/tunnels/{tid}/anchors")
    def anchors(tid: int):
        return service.list_anchors(tid)

    @app.put("/api/tunnels/{tid}/anchors/{seq}")
    async def put_anchor(tid: int, seq: int, body: AnchorBody):
        try:
            service.set_anchor(tid, seq, body.mileage_m)
        except (AnchorOrderError, AnchorRangeError) as e:
            raise HTTPException(400, str(e))
        hub.broadcast(tid, {"type": "anchor_update", "anchor": {"group_seq": seq, "mileage_m": body.mileage_m}})
        return {"ok": True}

    @app.delete("/api/tunnels/{tid}/anchors/{seq}")
    async def delete_anchor(tid: int, seq: int):
        try:
            service.delete_anchor(tid, seq)
        except KeyError:
            raise HTTPException(404, "錨點不存在")
        hub.broadcast(tid, {"type": "anchor_delete", "group_seq": seq})
        return {"ok": True}

    # ---------- 異狀類型（全工作區共用） ----------

    @app.get("/api/defect-types")
    def list_defect_types():
        return workspace.defect_types()

    @app.post("/api/defect-types")
    def add_defect_type(body: DefectTypeBody):
        try:
            return workspace.add_defect_type(body.name)
        except KeyError as e:
            raise HTTPException(409, str(e.args[0]))
        except ValueError as e:
            raise HTTPException(400, str(e))

    @app.delete("/api/defect-types/{type_id}")
    def remove_defect_type(type_id: int):
        try:
            action = workspace.remove_defect_type(type_id)
        except KeyError:
            raise HTTPException(404, "類型不存在")
        return {"action": action}

    # ---------- 專案歸檔（R9） ----------

    @app.get("/api/projects")
    def list_projects():
        return workspace.list_projects()

    @app.post("/api/projects")
    def create_project(body: ProjectBody):
        try:
            return workspace.create_project(body.name)
        except KeyError as e:
            raise HTTPException(409, str(e.args[0]))
        except ValueError as e:
            raise HTTPException(400, str(e))

    @app.put("/api/projects/{pid}")
    def rename_project(pid: int, body: ProjectBody):
        try:
            workspace.rename_project(pid, body.name)
        except KeyError as e:
            # Workspace 對「不存在」與「名稱撞名」皆丟 KeyError
            msg = str(e.args[0]) if e.args else ""
            raise HTTPException(409 if "已存在" in msg else 404, msg or "專案不存在")
        except ValueError as e:
            raise HTTPException(400, str(e))
        return {"ok": True}

    @app.delete("/api/projects/{pid}")
    def delete_project(pid: int):
        """刪除專案：底下隧道回到未分類（FK ON DELETE SET NULL），隧道不動。"""
        try:
            workspace.delete_project(pid)
        except KeyError:
            raise HTTPException(404, "專案不存在")
        return {"ok": True}

    @app.post("/api/tunnels/{tid}/move")
    def move_tunnel(tid: int, body: MoveBody):
        try:
            workspace.move_tunnel(tid, body.project_id)
        except KeyError:
            raise HTTPException(404, "隧道或專案不存在")
        return {"ok": True}

    # ---------- 照片標註（備註＋異狀） ----------

    @app.get("/api/tunnels/{tid}/photos/{pid}/annotation")
    def get_annotation(tid: int, pid: int):
        try:
            return service.annotation(tid, pid)
        except KeyError:
            raise HTTPException(404, "照片不存在")

    @app.put("/api/tunnels/{tid}/photos/{pid}/annotation")
    def put_annotation(tid: int, pid: int, body: AnnotationBody):
        try:
            result = service.set_annotation(tid, pid, body.note, [i.model_dump() for i in body.items])
        except KeyError:
            raise HTTPException(404, "照片不存在")
        except ValueError as e:
            raise HTTPException(400, str(e))
        # R9：備註/異狀存 DB 不改像素，不再失效縮圖快取
        hub.broadcast(
            tid,
            {
                "type": "annotation_updated",
                "photo_id": pid,
                "group_seq": result.get("group_seq"),
                "est_mileage_m": result.get("est_mileage_m"),
            },
        )
        return result

    @app.get("/api/tunnels/{tid}/anomalies")
    def anomalies_overview(
        tid: int,
        type_id: str = "",
        q: str = "",
        order: str = "asc",
    ):
        ids = []
        for part in type_id.split(","):
            part = part.strip()
            if part:
                try:
                    ids.append(int(part))
                except ValueError:
                    raise HTTPException(400, f"無效的 type_id：{part}")
        if order not in ("asc", "desc"):
            raise HTTPException(400, "order 僅接受 asc 或 desc")
        return service.anomaly_overview(tid, type_ids=ids or None, q=q or None, order=order)

    @app.get("/api/tunnels/{tid}/anomalies/export")
    def export_anomalies(
        tid: int,
        type_id: str = "",
        q: str = "",
        order: str = "asc",
        format: str = "xlsx",
    ):
        ids: list[int] = []
        for part in type_id.split(","):
            part = part.strip()
            if part:
                try:
                    ids.append(int(part))
                except ValueError:
                    raise HTTPException(400, f"無效的 type_id：{part}")
        if order not in ("asc", "desc"):
            raise HTTPException(400, "order 僅接受 asc 或 desc")
        if format not in ("csv", "xlsx"):
            raise HTTPException(400, "format 僅接受 csv 或 xlsx")
        rows = service.anomaly_overview(tid, type_ids=ids or None, q=q or None, order=order)
        try:
            meta = service.meta(tid)
            tunnel_name = meta.get("name", f"tunnel_{tid}")
        except Exception:
            tunnel_name = f"tunnel_{tid}"
        safe_name = "".join(c if c.isascii() and c.isalnum() or c in "_-" else "_" for c in tunnel_name)[:30].strip("_") or f"tunnel_{tid}"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if format == "csv":
            output = io.StringIO()
            output.write("\ufeff")
            writer = csv.writer(output)
            writer.writerow(["序號", "樁號", "里程(m)", "群組", "相機", "異狀類型", "異狀備註", "照片備註", "照片檔名", "相對路徑", "完整路徑", "建立時間"])
            for idx, r in enumerate(rows, 1):
                mileage = r.get("est_mileage_m")
                mileage_str = f"K{mileage//1000}+{mileage%1000:03d}" if isinstance(mileage, int) else ""
                rel = r.get("rel_path", "") or ""
                root = r.get("root_path", "") or ""
                full_path = str(Path(root) / rel) if root and rel else rel
                filename = Path(rel).name if rel else ""
                grp = r.get("group_seq")
                grp_display = grp + 1 if isinstance(grp, int) else ""
                writer.writerow([
                    idx,
                    mileage_str,
                    mileage if isinstance(mileage, int) else "",
                    grp_display,
                    r.get("camera_name", "") or "",
                    r.get("type_name", "") or "",
                    r.get("anomaly_note", "") or "",
                    r.get("photo_note", "") or "",
                    filename,
                    rel,
                    full_path,
                    r.get("created_at", "") or "",
                ])
            content = output.getvalue().encode("utf-8")
            filename = f"{safe_name}_anomalies_{timestamp}.csv"
            return Response(
                content=content,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
            )
        else:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                raise HTTPException(500, "未安裝 openpyxl，無法產生 xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "異狀清單"
            headers = ["序號", "樁號", "里程(m)", "群組", "相機", "異狀類型", "異狀備註", "照片備註", "照片檔名", "相對路徑", "完整路徑", "建立時間"]
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10, name="Microsoft JhengHei")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style="thin", color="B0B0B0")
            header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = header_border
            ws.row_dimensions[1].height = 22
            for idx, r in enumerate(rows, 1):
                mileage = r.get("est_mileage_m")
                mileage_str = f"K{mileage//1000}+{mileage%1000:03d}" if isinstance(mileage, int) else ""
                rel = r.get("rel_path", "") or ""
                root = r.get("root_path", "") or ""
                full_path = str(Path(root) / rel) if root and rel else rel
                filename = Path(rel).name if rel else ""
                grp = r.get("group_seq")
                grp_display = grp + 1 if isinstance(grp, int) else ""
                ws.append([
                    idx,
                    mileage_str,
                    mileage if isinstance(mileage, int) else "",
                    grp_display,
                    r.get("camera_name", "") or "",
                    r.get("type_name", "") or "",
                    r.get("anomaly_note", "") or "",
                    r.get("photo_note", "") or "",
                    filename,
                    rel,
                    full_path,
                    r.get("created_at", "") or "",
                ])
            data_font = Font(size=9, name="Microsoft JhengHei")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
                for cell in row:
                    cell.font = data_font
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    cell.border = header_border
            widths = [6, 10, 9, 7, 10, 12, 20, 20, 18, 30, 45, 19]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.print_title_rows = "1:1"
            ws.insert_rows(1)
            ws["A1"] = f"隧道：{tunnel_name}  |  異狀總數：{len(rows)}  |  匯出時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A1"].font = Font(bold=True, size=11, name="Microsoft JhengHei", color="1F4E78")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.row_dimensions[1].height = 18
            output = io.BytesIO()
            wb.save(output)
            content = output.getvalue()
            filename = f"{safe_name}_anomalies_{timestamp}.xlsx"
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
            )

    @app.get("/api/fs/list")
    def fs_list(path: str = ""):
        if not path.strip():
            return {
                "roots": platform_roots(),
                "recent": workspace.get_recent_paths(),
                "path": "",
                "parent": None,
                "dirs": [],
                "sample": None,
            }
        target = Path(path).expanduser()
        if not target.is_dir():
            raise HTTPException(404, "資料夾不存在")
        # scandir：is_dir/is_file 直接用列舉快取，避免 NAS/SMB 上每檔一次 stat 的網路往返
        dirs: list[str] = []
        sample: str | None = None
        with os.scandir(target) as it:
            for entry in it:
                try:
                    if entry.is_dir():
                        if not entry.name.startswith("."):
                            dirs.append(entry.name)
                        continue
                    if (
                        entry.is_file()
                        and Path(entry.name).suffix.lower() in {".jpg", ".jpeg"}
                        # scandir 順序任意，取檔名最小者為第一張（與舊 iterdir+sort 行為一致）
                        and (sample is None or entry.name.lower() < sample.lower())
                    ):
                        sample = entry.name
                except OSError:
                    continue
        dirs.sort(key=lambda n: n.lower())
        return {
            "path": str(target),
            "parent": str(target.parent) if target.parent != target else None,
            "dirs": dirs,
            "sample": sample,
            "recent": workspace.get_recent_paths(),
            "roots": platform_roots(),
        }

    @app.get("/api/fs/photo")
    def fs_photo(path: str, w: int | None = None):
        f = Path(path).expanduser()
        if not f.is_file() or f.suffix.lower() not in {".jpg", ".jpeg"}:
            raise HTTPException(404, "照片不存在")
        if w is None:
            return FileResponse(f, media_type="image/jpeg")
        # 縮圖快取（NAS 首張預覽加速）：w>0 時 draft+resize 並快取，與 /photos 共用策略但 key 為路徑雜湊
        try:
            w = int(w)
            if w <= 0 or w > 3000:
                raise ValueError()
        except Exception:
            raise HTTPException(400, "w 需為 1~3000")
        cache_dir = Path(workspace.root) / ".thumb_cache"
        # 以路徑雜湊避免非法檔名與過長路徑
        import hashlib

        key = hashlib.sha1(str(f.resolve()).encode()).hexdigest()[:16]
        needs_transpose = _needs_exif_transpose(f)
        cache = cache_dir / f"fs_{key}_{w}{'_t' if needs_transpose else ''}.jpg"
        if not cache.exists():
            img = Image.open(f)
            if needs_transpose:
                img = ImageOps.exif_transpose(img)
                img = img.convert("RGB")
            else:
                img.draft("RGB", (w * 2, w * 2))
                img = img.convert("RGB")
            ratio = w / img.width if img.width else 1
            img = img.resize((w, max(1, round(img.height * ratio))), Image.BILINEAR)
            cache_dir.mkdir(parents=True, exist_ok=True)
            img.save(cache, "JPEG", quality=87)
        return FileResponse(cache, media_type="image/jpeg")

    @app.get("/api/tunnels/{tid}/info")
    def tunnel_info(tid: int):
        # R9：進入檢視器＝最近使用時間觸碰點
        try:
            workspace.touch_tunnel(tid)
        except Exception:
            pass
        return service.info(tid)

    @app.post("/api/tunnels/{tid}/photos/{pid}/mark_missing")
    def mark_missing(tid: int, pid: int):
        try:
            service.mark_missing_photo(tid, pid)
        except KeyError:
            raise HTTPException(404, "照片不存在")
        # R9：改判缺照不改變像素，不失效縮圖
        hub.broadcast(tid, {"type": "photo_updated", "photo_id": pid})
        return {"ok": True}

    @app.post("/api/tunnels/{tid}/photos/{pid}/restore")
    def restore_missing(tid: int, pid: int):
        try:
            service.restore_missing_photo(tid, pid)
        except KeyError:
            raise HTTPException(404, "照片不存在或未改判")
        hub.broadcast(tid, {"type": "photo_updated", "photo_id": pid})
        return {"ok": True}

    @app.post("/api/tunnels/{tid}/realign")
    def realign_dry_run(tid: int, body: RealignBody):
        return service.realign_preview(tid, body.tolerance_seconds)

    @app.post("/api/tunnels/{tid}/realign/apply")
    async def realign_apply(tid: int, body: RealignBody):
        result = service.realign_apply(tid, body.tolerance_seconds)
        # R9：重新對齊只改群組歸屬，不改像素，不失效縮圖
        hub.broadcast(tid, {"type": "realigned"})
        return result

    @app.post("/api/tunnels/{tid}/groups/{seq}/merge")
    async def merge_group(tid: int, seq: int, body: MergeBody):
        try:
            result = service.merge_group(tid, seq, body.direction, body.keep)
        except MergeConflict as e:
            raise HTTPException(409, detail={"message": str(e), "conflict_cameras": e.conflict_cameras})
        except KeyError:
            raise HTTPException(404, "群組不存在")
        hub.broadcast(tid, {"type": "merged", **result})
        return result

    @app.put("/api/tunnels/{tid}/cameras/{seq}")
    def update_camera(tid: int, seq: int, body: CameraUpdateBody):
        if body.name is None and body.rotation is None and body.grid_pos is None:
            raise HTTPException(400, "需提供 name、rotation 或 grid_pos")
        if body.rotation is not None and (
            body.rotation % 90 != 0 or not (0 <= body.rotation <= 270)
        ):
            raise HTTPException(400, "旋轉角度僅接受 0/90/180/270")
        try:
            if body.name is not None:
                service.set_camera_name(tid, seq, body.name)
            if body.rotation is not None:
                service.set_camera_rotation(tid, seq, body.rotation)
            if body.grid_pos is not None:
                service.set_camera_grid_pos(tid, seq, body.grid_pos)
        except KeyError:
            raise HTTPException(404, "相機不存在")
        except ValueError as e:
            raise HTTPException(400, str(e))
        # R9 精準失效：僅旋轉改變像素 → 遞增該機位全部照片的 pixel_version 並清其縮圖
        if body.rotation is not None:
            _bump_pixels(service, workspace, tid, service.camera_photo_ids(tid, seq))
        hub.broadcast(tid, {"type": "camera_updated", "camera_seq": seq})
        return {"ok": True}

    @app.put("/api/tunnels/{tid}/layout")
    def set_layout(tid: int, body: LayoutBody):
        try:
            service.set_layout_cols(tid, body.cols)
        except ValueError as e:
            raise HTTPException(400, str(e))
        hub.broadcast(tid, {"type": "layout_updated"})
        return {"ok": True}

    @app.post("/api/fs/recent")
    def record_recent(body: FsRecentBody):
        target = Path(body.path).expanduser()
        if not target.is_dir():
            raise HTTPException(404, "資料夾不存在")
        workspace.add_recent_path(str(target))
        return {"ok": True}

    @app.put("/api/tunnels/{tid}/photos/{pid}/rotation")
    def set_photo_rotation(tid: int, pid: int, body: PhotoRotationBody):
        if body.angle % 90 != 0 or not (0 <= body.angle <= 270):
            raise HTTPException(400, "旋轉角度僅接受 0/90/180/270")
        try:
            service.set_photo_rotation(tid, pid, body.angle)
        except KeyError:
            raise HTTPException(404, "照片不存在")
        # R9 精準失效：單張像素版本遞增＋清該張縮圖
        _bump_pixels(service, workspace, tid, [pid])
        hub.broadcast(tid, {"type": "photo_updated", "photo_id": pid})
        return {"ok": True}

    @app.get("/api/tunnels/{tid}/orientation-stats")
    def get_orientation_stats(tid: int):
        return service.orientation_stats(tid)

    @app.post("/api/tunnels/{tid}/cameras/{seq}/unify")
    def unify_camera_orientation(tid: int, seq: int, body: PhotoRotationBody):
        # 批次轉正：只接受 90/270（180 不改變直橫、0 無意義）
        if body.angle not in (90, 270):
            raise HTTPException(400, "統一方向僅接受 90 或 270")
        try:
            updated = service.unify_camera_orientation(tid, seq, body.angle)
        except KeyError:
            raise HTTPException(404, "相機不存在")
        # R9 精準失效：批次轉正改變該機位像素 → 遞增版本並清縮圖
        _bump_pixels(service, workspace, tid, service.camera_photo_ids(tid, seq))
        hub.broadcast(tid, {"type": "camera_updated", "camera_seq": seq})
        return {"ok": True, "updated": updated}

    @app.get("/api/tunnels/{tid}/camera_thumbs")
    def camera_thumbs(tid: int):
        return service.camera_thumbs(tid)

    @app.get("/api/tunnels/{tid}/photos/{photo_id}")
    def photo(tid: int, photo_id: int, w: int | None = None, pv: int | None = None, request: Request = None):
        inm = _request_etags(request) if request is not None else set()
        try:
            info = service.photo_render_info(tid, photo_id)
        except KeyError:
            raise HTTPException(404, "照片不存在")
        path: Path = info["path"]
        extra: int = info["extra_rotation"]
        # 快取檔版本取 URL 傳入的 pv（前端一律帶 groups API 給的當前值）；
        # 未帶時退回 DB 當前值。bump 時會清除該照片所有版本的快取檔。
        pv_val = pv if pv is not None else int(info.get("pixel_version") or 0)
        if not Path(path).exists():
            raise HTTPException(404, "照片檔案遺失（原檔可能被移動）")

        # R9：orientation 入庫——DB 為 NULL（舊隧道）時 lazy backfill 一次
        orientation = info.get("orientation")
        if orientation is None:
            orientation = _read_orientation_tag(path)
            try:
                service.set_photo_orientation(tid, photo_id, orientation)
            except Exception:
                pass
        needs_transpose = int(orientation) not in (0, 1)

        fast_path = w is None and extra == 0 and not needs_transpose
        if fast_path:
            # 原檔直出：非 immutable（原檔可能被外部取代），ETag 協商＋條件請求 304
            etag = f'"{path.stat().st_mtime_ns:x}-{path.stat().st_size:x}"'
            if inm and etag in inm:
                return Response(status_code=304, headers={"ETag": etag, "Cache-Control": "public, max-age=3600"})
            return FileResponse(
                path,
                media_type="image/jpeg",
                headers={"Cache-Control": "public, max-age=3600", "ETag": etag},
            )

        suffix = "orig" if w is None else str(w)
        quality = thumbs.ORIG_QUALITY if w is None else thumbs.THUMB_QUALITY
        cache_dir = Path(workspace.root) / ".thumb_cache"
        cache = cache_dir / f"{tid}_{photo_id}_{suffix}_{extra}_v{pv_val}.jpg"

        def _produce():
            return thumbs.make_thumbnail(
                path,
                w,
                needs_transpose=needs_transpose,
                extra_rotation=extra,
                quality=quality,
            )

        try:
            thumbs.get_or_make(cache, _produce)  # R9：single-flight，併發不重複解碼
        except RuntimeError:
            raise HTTPException(500, "縮圖生成失敗")
        # 內容版本已編入 URL（pv）→ immutable 安全；仍附 ETag 供明確協商
        thumb_etag = f'"{cache.stat().st_mtime_ns:x}-{cache.stat().st_size:x}"'
        if inm and thumb_etag in inm:
            return Response(status_code=304, headers={"ETag": thumb_etag, "Cache-Control": "public, max-age=31536000, immutable"})
        return FileResponse(
            cache,
            media_type="image/jpeg",
            headers={
                "Cache-Control": "public, max-age=31536000, immutable",
                "ETag": thumb_etag,
            },
        )

    @app.websocket("/ws/tunnels/{tid}")
    async def ws_room(websocket: WebSocket, tid: int):
        await hub.join(tid, websocket)
        try:
            while True:
                await websocket.receive_text()
        except WebSocketDisconnect:
            hub.leave(tid, websocket)

    dist = _find_dist()
    if dist is not None:
        app.mount("/", StaticFiles(directory=str(dist), html=True), name="web")

    return app


def _find_dist():
    env = os.environ.get("TUNNELVIEW_DIST")
    if env and Path(env).is_dir():
        return Path(env)
    candidate = Path(__file__).resolve().parents[2] / "frontend" / "dist"
    if candidate.is_dir():
        return candidate
    return None
